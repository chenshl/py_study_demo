#!/usr/bin/python
# coding:utf-8
# @author : csl
# @date   : 2018/03/26 22:02
'''Excel表格操作'''

# 读写.xls表格03-07版
import xlrd
import xlwt
# 读写.xlsx表格07版
import openpyxl

def write_03_Excel(file_path):
    wb = xlwt.Workbook()  #打开excel文件
    sheet = wb.add_sheet("测试表格2003")  #添加表格名称
    value = [["姓名", "年龄", "电话", "婚姻状况"],
             ["范彬彬", "22", "18888888888", "已婚"],
             ["袁姗姗", "25", "18999999999", "未婚"],
             ["刘德华", "50", "17777777777", "已婚"],
             ["张学友", "55", "15555555555", "已婚"],
             ["郭富城", "55", "13333333333", "已婚"]]
    for i in range(0, len(value)):
        for e in range(0, len(value[i])):
            sheet.write(i, e, value[i][e])  #i、e分别表示行和列
    wb.save(file_path)
    print("写入表格成功！")


def read_03_Excel(file_path):
    wb = xlrd.open_workbook(file_path)  #打开excel文件
    r_sheet = wb.sheet_names()  #查找所有的表名
    work_sheet = wb.sheet_by_name(r_sheet[0])  #通过表名找到第一张表
    for i in range(0, work_sheet.nrows):  #循环所有行
        row = work_sheet.row(i)  #获取第i行
        for j in range(0, work_sheet.ncols):  #循环所有的列
            print(work_sheet.cell_value(i, j), "\t", end="")  #获取行和列的cell
        print()


def wirte_07_Excel(file_path):
    wb = openpyxl.Workbook()  #打开文件
    sheet = wb.active  #激活sheet表格
    sheet.title = "测试表格2007"  #添加sheet表格名称
    value = [["姓名", "年龄", "电话", "婚姻状况"],
             ["范彬彬", "22", "18888888888", "已婚"],
             ["袁姗姗", "25", "18999999999", "未婚"],
             ["刘德华", "50", "17777777777", "已婚"],
             ["张学友", "55", "15555555555", "已婚"],
             ["郭富城", "55", "13333333333", "已婚"]]
    for i in range(0, len(value)):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))  #写入单元格
    wb.save(file_path)
    print("写入07表格成功")


def read_07_Excel(file_path):
    wb = openpyxl.load_workbook(file_path)  #打开文件
    # sheet = wb.get_sheet_by_name("测试表格2007")
    sheet = wb["测试表格2007"]  #通过sheet名称锁定表格
    for row in sheet.rows:  #循环所有的行
        for cell in row:  #循环行中所有的单元格
            print(cell.value, "\t", end="")  #获取单元格的值
        print()

file_03_excel = "./data/03excel.xls"
file_07_excel = "./data/07excel.xlsx"
write_03_Excel(file_03_excel)
read_03_Excel(file_03_excel)
wirte_07_Excel(file_07_excel)
read_07_Excel(file_07_excel)
