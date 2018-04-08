#!/usr/bin/python
# coding:utf-8
# @author : csl
# @date   : 2018/04/05 22:12
# 通过Excel执行的接口测试用例报告以Excel形式生成
# 可以先查询到结果表格的行数，再在第几行接着写入

import openpyxl
import time

class result_datas(object):

    def __init__(self, api_name, code, sheet_name="测试报告", result_data=None):
        self.api_name = api_name
        self.code = code
        self.sheet_name = sheet_name
        self.result_data = result_data
        # ctime = time.strftime(time.localtime(time.time()), "%Y%m%d")
        repath = "../test_datas/"
        # self.re_name = repath + "result_datas" + ctime + ".xlsx"
        self.re_name = repath + "result_datas" + ".xlsx"

    def write_result_2_Excel(self):
        v_col = ["接口名称", "响应状态码", "响应结果"]
        # 打开文件
        r_excel = openpyxl.load_workbook(self.re_name)
        r_sheet = r_excel.active  # 激活sheet表格
        # print(r_sheet)
        r_sheet.title = self.sheet_name
        rows = r_sheet.max_row  # 获取已存在行数
        cols = r_sheet.max_column  # 获取已存在的列数
        if rows <= 1:
            for c in list(range(len(v_col))):
                r_sheet.cell(row=1, column=c+1, value=str(v_col[c]))
        r_sheet.cell(row=rows+1, column=1, value=self.api_name)  # 在下一行写入
        r_sheet.cell(row=rows+1, column=2, value=self.code)
        r_sheet.cell(row=rows+1, column=3, value=self.result_data)
        r_excel.save(self.re_name)
        print("写入结果完成")


if __name__ == "__main__":
    api_name = "/login"
    code = "200"
    result_data = "aaa"
    result_datas(api_name, code, "测试报告1", result_data).write_result_2_Excel()
